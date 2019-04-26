from openpyxl import load_workbook
#B4 ~ X24
wb = load_workbook(filename = 'test.xlsx')
sheet1 = wb['Sheet1']
sheet2 = wb.active

allData = []
#for row in sheet1.iter_rows(min_row=
#print(sheet1['D8'].value)
for i in sheet1.iter_rows(min_row=4, max_col=24, max_row=23):
	for j in i:
		if j.value in isdigit():
		print(j.value,"\n")
